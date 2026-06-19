import hashlib
import io
import json
import os
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from core.data_loader import load_dashboard_data

ALL_METRICS = ["할인율", "BEST상품", "신선도", "시즌"]


class _ExcelStyles:
    def __init__(self):
        self.align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.border_thin = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self.fill_header = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        self.fill_green = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        self.fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        self.fill_amt = self.fill_green
        self.font_bold = Font(name='맑은 고딕', bold=True, size=9)
        self.font_title = Font(name='맑은 고딕', bold=True, size=10)
        self.font_white_bold = Font(name='맑은 고딕', bold=True, color="FFFFFF", size=9)
        self.font_black = Font(name='맑은 고딕', color="000000", size=9)

    def fill_for_score(self, actual_score, target_score):
        if target_score <= 0:
            return self.fill_yellow
        ratio = actual_score / target_score
        if ratio > 1.2:
            return self.fill_yellow
        if ratio >= 1.0:
            return self.fill_green
        if ratio >= 0.7:
            return self.fill_yellow
        return self.fill_red

    def font_for_fill(self, fill):
        return self.font_black if fill is self.fill_yellow else self.font_white_bold

    def apply_score_cell(self, cell, value, actual, target):
        cell.value = value
        fill = self.fill_for_score(actual, target)
        cell.fill = fill
        cell.font = self.font_for_fill(fill)
        cell.alignment = self.align_center
        cell.border = self.border_thin

    def apply_plain_cell(self, cell, value):
        cell.value = value
        cell.alignment = self.align_center
        cell.border = self.border_thin

    def apply_yellow_cell(self, cell, value):
        cell.value = value
        cell.fill = self.fill_yellow
        cell.font = self.font_black
        cell.alignment = self.align_center
        cell.border = self.border_thin

    def apply_header_cell(self, cell, value=None):
        if value is not None:
            cell.value = value
        cell.font = self.font_bold
        cell.alignment = self.align_center
        cell.border = self.border_thin
        cell.fill = self.fill_header


_STYLES = _ExcelStyles()


def _round1(n):
    """소수점 둘째 자리에서 반올림 → 첫째 자리(0.0)"""
    return round(float(n or 0), 1)


def _fmt_num(n):
    val = _round1(n)
    if val == 0: return "0"
    return f"{val:.1f}"


def _fmt_score(n):
    val = _round1(n)
    if val == 0: return "0점"
    return f"{val:.1f}점"


def _normalize_metrics_filter(metrics_filter):
    if not metrics_filter:
        return list(ALL_METRICS)
    out = []
    for raw in metrics_filter:
        m = str(raw).strip()
        if m in ("MD구성", "BEST", "BEST상품"):
            key = "BEST상품"
        elif m in ALL_METRICS:
            key = m
        else:
            continue
        if key not in out:
            out.append(key)
    return out or list(ALL_METRICS)


def _extract_seg_keys_labels(b_detail, m_id, fallback_keys, fallback_labels):
    LABEL_MAP = {
        "d70": "70% 이상", "d50": "50~70% 미만", "d30": "30~50% 미만", "d10": "1~30% 미만", "d0": "0%",
        "new": "신상", "plan": "기획", "spring": "봄 (SS)", "summer": "여름 (SS)",
        "Outer": "아우터", "Top": "상의", "Bottom": "하의", "Skirt": "스커트", "Dress": "원피스",
        "best": "판매 TOP10", "70%이상": "70% 이상", "50-69%": "50~70% 미만", "30-49%": "30~50% 미만",
        "1-29%": "1~30% 미만", "정상가": "0%", "봄(SS)": "봄 (SS)", "여름(SS)": "여름 (SS)"
    }
    if not b_detail or not isinstance(b_detail, dict):
        return fallback_keys, [LABEL_MAP.get(x, x) for x in fallback_labels]
    m_data = b_detail.get(m_id)
    if not m_data or not isinstance(m_data, dict):
        return fallback_keys, [LABEL_MAP.get(x, x) for x in fallback_labels]
    segs = m_data.get("segs", [])
    if not segs:
        return fallback_keys, [LABEL_MAP.get(x, x) for x in fallback_labels]
    # 시즌의 경우 is_score_target=True(현 시즌) 세그먼트만 포함
    if m_id == "season":
        segs = [s for s in segs if isinstance(s, dict) and s.get("is_score_target", False)]
    keys = [s["key"] for s in segs if isinstance(s, dict) and "key" in s]
    raw_labels = [s.get("l") or s.get("label") or s["key"] for s in segs if isinstance(s, dict) and "key" in s]
    labels = [LABEL_MAP.get(lbl, lbl) for lbl in raw_labels]
    return (keys or fallback_keys), (labels or [LABEL_MAP.get(x, x) for x in fallback_labels])


def evaluate_color(actual_score, target_score):
    fill = _STYLES.fill_for_score(actual_score, target_score)
    if fill is _STYLES.fill_green:
        return "00B050"
    if fill is _STYLES.fill_red:
        return "FF0000"
    return "FFFF00"


def dashboard_fingerprint(data: dict) -> str:
    sig = json.dumps(
        {
            "n": len(data.get("BRANDS", [])),
            "stores": data.get("STORES"),
            "cats": data.get("CATS"),
            "brands": [
                (
                    b.get("store"),
                    b.get("name"),
                    b.get("category"),
                    b.get("total"),
                    b.get("dis"),
                    b.get("fresh"),
                    b.get("best"),
                    b.get("season"),
                    b.get("item"),
                )
                for b in data.get("BRANDS", [])
            ],
        },
        ensure_ascii=False,
        sort_keys=True,
        default=str,
    )
    return hashlib.md5(sig.encode("utf-8")).hexdigest()


def _sanitize_sheet_name(store: str, used: set) -> str:
    name = re.sub(r"[\\/*?:\[\]]", "_", str(store or "지점").strip())
    name = name[:31] if name else "지점"
    base, n = name, 2
    while name in used:
        suffix = f"_{n}"
        name = f"{base[: 31 - len(suffix)]}{suffix}"
        n += 1
    used.add(name)
    return name


def _sheet_title_from_filters(store_filter: str, cat_filter: str) -> str:
    """P4: 시트 탭명 = [지점명]_[카테고리명]"""
    raw = f"{store_filter}_{cat_filter}"
    return _sanitize_sheet_name(raw, set())


def _fill_exposure_sheet(
    ws,
    filtered_brands,
    detail_data,
    title_store: str,
    title_cat: str,
    metrics_filter,
    show_store_col: bool = False,
    hide_brand_col: bool = False,
    score_mode: str = "weighted",
    is_sales_execution: bool = False,
):
    styles = _STYLES
    if not filtered_brands:
        return

    metrics_filter = _normalize_metrics_filter(metrics_filter)

    # ── 1단계: 각 브랜드의 모든 세그먼트 점수를 사전 계산하여 120% 감점 적용 및 정확한 총점수(calculated_total)를 역산 ──
    for b in filtered_brands:
        if b.get("is_p1_aggregate"):
            # P1 집계 브랜드의 경우, calculated_total은 이미 정확하게 계산되어 있으므로 덮어쓰지 않고 넘어감
            continue
        b_store = b.get("store", "")
        b_name = b.get("name", "")
        b_type = b.get("type_label", "")
        b_store_dict = detail_data.get(b_store) or {}
        b_name_dict = b_store_dict.get(b_name) or {}
        b_detail = b_name_dict.get(b_type) or {}
        s_weights = b.get("scoring_guide", {}).get("score_weights", {})
        
        b_earned_scores = {}
        b_max_weights = {}
        for m_id in ("dis", "fresh", "best", "season", "item"):
            m_data = b_detail.get(m_id) or {}
            w_key = "sea" if m_id == "season" else m_id
            m_weight = s_weights.get(w_key, 0.0)
            
            segs_raw = m_data.get("segs") if isinstance(m_data, dict) else None
            if segs_raw:
                # 세그먼트 opt_pct 비중 합계 수집
                seg_opt_pct_map = {}
                sum_opt_pct = 0.0
                for s in segs_raw:
                    if isinstance(s, dict) and "key" in s:
                        op_val = float(s.get("opt_pct", 0.0))
                        seg_opt_pct_map[s["key"]] = op_val
                        sum_opt_pct += op_val
                if sum_opt_pct <= 0:
                    sum_opt_pct = 100.0

                sum_earned_pt = 0.0
                for s in segs_raw:
                    if not isinstance(s, dict) or "key" not in s:
                        continue
                    key = s["key"]
                    brand_opt_pct = seg_opt_pct_map.get(key, 0.0)
                    # 정규화된 만점 적용
                    brand_seg_max_pt = m_weight * (brand_opt_pct / sum_opt_pct)
                    
                    valM = float(s.get("valM", 0.0))
                    targetM = float(s.get("targetM", 0.0))
                    pct = float(s.get("pct", 0.0))
                    
                    real_pct = (valM / targetM * 100.0) if targetM > 0 else pct
                    is_over_120 = (real_pct > 120.0)
                    
                    # pct 를 최대 100으로 제한한 후 만점에 적용 (만점 초과 방지)
                    capped_pct = min(pct, 100.0)
                    raw_earned_pt = brand_seg_max_pt * (capped_pct / 100.0)
                    earned_pt = raw_earned_pt * 0.9 if is_over_120 else raw_earned_pt
                    # 세그먼트 만점을 초과하지 않도록 추가 켜핑
                    earned_pt = min(earned_pt, brand_seg_max_pt)
                    sum_earned_pt += earned_pt
                
                raw_earned = min(sum_earned_pt, m_weight)
                if score_mode == "100_percent" and m_weight > 0:
                    b_earned_scores[m_id] = (raw_earned / m_weight) * 100.0
                else:
                    b_earned_scores[m_id] = raw_earned
            else:
                raw_score_0_to_100 = float(b.get(m_id, 0.0))
                if score_mode == "100_percent":
                    b_earned_scores[m_id] = raw_score_0_to_100
                else:
                    b_earned_scores[m_id] = raw_score_0_to_100 * (m_weight / 100.0)
            
            if score_mode == "100_percent":
                b_max_weights[m_id] = 100.0 if m_weight > 0 else 0.0
            else:
                b_max_weights[m_id] = m_weight

        metric_filter_map = {"dis": "할인율", "best": "BEST상품", "fresh": "신선도", "season": "시즌", "item": "아이템"}
        is_all = ("전체" in metrics_filter) or (len(metrics_filter) >= 4)

        tot = 0.0
        max_tot = 0.0
        for m_id in ("dis", "fresh", "best", "season", "item"):
            if is_all or metric_filter_map[m_id] in metrics_filter:
                tot += b_earned_scores[m_id]
                max_tot += b_max_weights.get(m_id, 0.0)

        if score_mode == "100_percent":
            # 100점 환산 기준인 경우 선택된 지표들의 단순 평균 점수가 100점 만점이 됨
            selected_count = sum(1 for m_id in ("dis", "fresh", "best", "season", "item") if (is_all or metric_filter_map[m_id] in metrics_filter) and b_max_weights.get(m_id, 0.0) > 0)
            if selected_count > 0:
                normalized_tot = tot / selected_count
            else:
                normalized_tot = 0.0
        else:
            if max_tot > 0 and max_tot < 100.0 and not is_all:
                normalized_tot = (tot / max_tot) * 100.0
            else:
                normalized_tot = tot

        b["calculated_total"] = _round1(min(normalized_tot, 100.0))

    sorted_brands = sorted(filtered_brands, key=lambda x: x.get("calculated_total", 0.0), reverse=True)

    # ── 2단계: 1위 브랜드 기준(sorted_brands[0])으로 메타데이터 및 지표 세그먼트 구성 정의 ──
    sample_b = sorted_brands[0]
    s_store = sample_b.get("store", "")
    s_name = sample_b.get("name", "")
    s_type = sample_b.get("type_label", "")
    
    s_store_dict = detail_data.get(s_store) or {}
    s_name_dict = s_store_dict.get(s_name) or {}
    sample_detail = s_name_dict.get(s_type) or {}

    item_seg_keys, item_labels = _extract_seg_keys_labels(
        sample_detail, "item",
        ["Outer", "Top", "Bottom", "Skirt", "Dress"],
        ["아우터", "상의", "하의", "스커트", "원피스"],
    )
    dis_keys, dis_labels = _extract_seg_keys_labels(
        sample_detail, "dis",
        ["d70", "d50", "d30", "d10", "d0"],
        ["70%이상", "50-69%", "30-49%", "1-29%", "정상가"],
    )
    fresh_keys, fresh_labels = _extract_seg_keys_labels(
        sample_detail, "fresh", ["new", "plan"], ["신상", "기획"]
    )
    season_keys, season_labels = _extract_seg_keys_labels(
        sample_detail, "season",
        ["spring", "summer"],  # SS 시즌 기본값 (FW: autumn/winter 제외)
        ["봄(SS)", "여름(SS)"],
    )

    # ── 3개 층 헤더 정의 ──
    ROW_TITLE, ROW_H1, ROW_H2, ROW_H3, ROW_DATA = 1, 2, 3, 4, 5

    # 공통 헤더 (랭크, 복종, 지점, 브랜드, 총점수, 총재고)
    common_headers = [
        ("랭크", 1), ("복종", 1)
    ]
    
    if show_store_col:
        common_headers.append(("지점", 1))

    if not hide_brand_col:
        common_headers.append(("브랜드", 1))

    common_headers.extend([("총\n점수\n(100점)", 1), ("총보유\n재고액", 1)])

    metrics_config = {}
    if "할인율" in metrics_filter:
        metrics_config["dis"] = {"title": "할인율", "keys": dis_keys, "labels": dis_labels}
    if "BEST상품" in metrics_filter:
        metrics_config["best"] = {"title": "BEST상품", "keys": ["best"], "labels": ["판매\nTOP10"]}
    if "신선도" in metrics_filter:
        metrics_config["fresh"] = {"title": "신선도", "keys": fresh_keys, "labels": fresh_labels}
    if "시즌" in metrics_filter:
        metrics_config["season"] = {"title": "시즌", "keys": season_keys, "labels": season_labels}
    if "아이템" in metrics_filter:
        metrics_config["item"] = {"title": "아이템", "keys": item_seg_keys, "labels": item_labels}

    for m_id, m_info in metrics_config.items():
        for extra_b in filtered_brands:
            eb_store = extra_b.get("store", "")
            eb_name = extra_b.get("name", "")
            eb_type = extra_b.get("type_label", "")
            eb_store_dict = detail_data.get(eb_store) or {}
            eb_name_dict = eb_store_dict.get(eb_name) or {}
            eb_detail = eb_name_dict.get(eb_type) or {}
            nk, nl = _extract_seg_keys_labels(eb_detail, m_id, m_info["keys"], m_info["labels"])
            if len(nk) > len(m_info["keys"]):
                m_info["keys"], m_info["labels"] = nk, nl

    total_cols = sum(span for _, span in common_headers) + sum(len(m_info["keys"]) + 1 for m_info in metrics_config.values())
    
    clean_store = title_store.replace("NC", "").replace("점", "").strip()
    clean_cat = title_cat.replace("카테고리", "").strip()
    sheet_title = f"[한국유통] 상품구색 노출/측정 ({clean_store} {clean_cat})"

    title_cell = ws.cell(row=ROW_TITLE, column=1, value=sheet_title)
    title_cell.font = styles.font_title
    title_cell.alignment = styles.align_center
    ws.merge_cells(start_row=ROW_TITLE, start_column=1, end_row=ROW_TITLE, end_column=total_cols)

    # ── 공통 헤더 3개 행 수직 병합 및 기재 ──
    col_idx = 1
    for hdr_title, colspan in common_headers:
        ws.cell(row=ROW_H1, column=col_idx, value=hdr_title)
        ws.merge_cells(
            start_row=ROW_H1, start_column=col_idx,
            end_row=ROW_H3, end_column=col_idx
        )
        col_idx += colspan

    # ── 지표 헤더 3층 기재 ──
    from config.scoring_config import get_weights_by_category
    
    # 각 지표별 정상/상설 만점을 저장해 두고 ROW_H3 세그먼트 컬럼에도 사용
    metric_norm_outl_w = {}  # {m_id: (norm_w, outl_w)}
    
    for m_id, m_info in metrics_config.items():
        colspan = len(m_info["keys"]) + 1

        cat = sample_b.get("category", "여성")
        b_name = sample_b.get("name", "")
        norm_cfg = get_weights_by_category(cat, "정상", b_name)
        outl_cfg = get_weights_by_category(cat, "상설", b_name)
        
        weight_key_map = {
            "dis": "weight_discount",
            "fresh": "weight_freshness",
            "season": "weight_season",
            "best": "weight_best",
            "item": "weight_item"
        }
        w_key = weight_key_map.get(m_id, "")
        norm_w = int(round(norm_cfg.get(w_key, 0.0) * 100)) if w_key else 0
        outl_w = int(round(outl_cfg.get(w_key, 0.0) * 100)) if w_key else 0
        metric_norm_outl_w[m_id] = (norm_w, outl_w)
        
        # H1 (2행): 지표명 + 상설/정상 점수를 한 줄로 표기
        # 정상≠상설인 경우: "할인율\n상설40점 / 정상30점"
        # 동일한 경우:       "할인율\n(30점)"
        if norm_w != outl_w and norm_w > 0 and outl_w > 0:
            header_val = f"{m_info['title']}\n상설{outl_w}점 / 정상{norm_w}점"
        elif norm_w > 0:
            header_val = f"{m_info['title']}\n({norm_w}점)"
        else:
            sample_w_map = sample_b.get("scoring_guide", {}).get("score_weights", {})
            sample_w_key = "sea" if m_id == "season" else m_id
            sample_w = sample_w_map.get(sample_w_key, 0)
            header_val = f"{m_info['title']}\n({int(round(sample_w))}점)"

        cell_h1 = ws.cell(row=ROW_H1, column=col_idx, value=header_val)
        cell_h1.alignment = styles.align_center
        ws.merge_cells(
            start_row=ROW_H1, start_column=col_idx,
            end_row=ROW_H1, end_column=col_idx + colspan - 1
        )

        # H2 (3행): 세그먼트 레이블 ("합계", "70%이상", ...)
        ws.cell(row=ROW_H2, column=col_idx, value="합계")
        for i, lbl in enumerate(m_info["labels"]):
            ws.cell(row=ROW_H2, column=col_idx + 1 + i, value=lbl)

        col_idx += colspan

    if is_sales_execution:
        total_cols += 1
        ws.cell(row=ROW_H1, column=total_cols, value="개선필요항목")
        ws.merge_cells(start_row=ROW_H1, start_column=total_cols, end_row=ROW_H3, end_column=total_cols)
        ws.column_dimensions[get_column_letter(total_cols)].width = 25

    # H1, H2, H3 헤더 스타일 일괄 주입
    for r in (ROW_H1, ROW_H2, ROW_H3):
        for c in range(1, total_cols + 1):
            styles.apply_header_cell(ws.cell(row=r, column=c))

    ws.row_dimensions[ROW_TITLE].height = 25
    ws.row_dimensions[ROW_H1].height = 30  # 2줄(지표명 + 점수) 표기
    ws.row_dimensions[ROW_H2].height = 20
    ws.row_dimensions[ROW_H3].height = 25  # 정상/상설 2줄 표기 위해 높이 증가

    empty_txt = "-\n(-)"
    row_idx = ROW_DATA

    ref_b = sorted_brands[0] if sorted_brands else None
    ref_weights = ref_b.get("scoring_guide", {}).get("score_weights", {}) if ref_b else {}

    for rank, b in enumerate(sorted_brands, 1):
        store = b.get("store", "")
        b_name = b.get("name", "")
        b_type = b.get("type_label", "")
        b_store_dict = detail_data.get(store) or {}
        b_name_dict = b_store_dict.get(b_name) or {}
        b_detail = b_name_dict.get(b_type) or {}
        s_weights = b.get("scoring_guide", {}).get("score_weights", {})

        col = 1
        styles.apply_plain_cell(ws.cell(row=row_idx, column=col), rank)
        col += 1
        
        styles.apply_plain_cell(ws.cell(row=row_idx, column=col), b.get("category", ""))
        col += 1

        if show_store_col:
            styles.apply_plain_cell(ws.cell(row=row_idx, column=col), store)
            col += 1

        if not hide_brand_col:
            styles.apply_plain_cell(ws.cell(row=row_idx, column=col), b_name)
            col += 1

        tot_score = b.get("calculated_total", 0.0)
        # P2 화면 표시: 각 지표 earned_score 합산으로 정확한 총점 표시 (is_p1_aggregate는 이미 calculated_total에 합산 완료)
        tot_score_fmt = f"{int(round(tot_score))}점"
        styles.apply_score_cell(
            ws.cell(row=row_idx, column=col), tot_score_fmt, tot_score, 100.0
        )
        col += 1

        s_val = float(b.get("sM", 0.0))
        t_val = float(b.get("tM", 0.0))
        c_amt = ws.cell(row=row_idx, column=col)
        c_amt.value = _fmt_num(s_val)
        
        # [v171] 총재고액 목표재고 대비 달성률 색상 판정 로직: 100% ~ 120% 이내만 초록색, 그 외는 빨간색 하이라이트
        target_stock = t_val * 2.0
        if target_stock > 0:
            stock_ratio = s_val / target_stock
            if 1.0 <= stock_ratio <= 1.2:
                c_amt.fill = styles.fill_green
                c_amt.font = styles.font_white_bold
            else:
                c_amt.fill = styles.fill_red
                c_amt.font = styles.font_white_bold

        c_amt.alignment = styles.align_center
        c_amt.border = styles.border_thin
        col += 1

        col_offset = col
        red_lights = []

        for m_id, m_info in metrics_config.items():
            m_data = b_detail.get(m_id) or {}
            w_key = "sea" if m_id == "season" else m_id
            m_weight = s_weights.get(w_key, 0.0)
            global_m_weight = ref_weights.get(w_key, m_weight)
            seg_keys = m_info["keys"]
            segs_raw = m_data.get("segs") if isinstance(m_data, dict) else None

            # 세그먼트별 opt_pct 및 합계 수집
            seg_opt_pct_map = {}
            sum_opt_pct = 0.0
            if segs_raw:
                for s in segs_raw:
                    if isinstance(s, dict) and "key" in s:
                        op_val = float(s.get("opt_pct", 0.0))
                        seg_opt_pct_map[s["key"]] = op_val
                        sum_opt_pct += op_val
            if sum_opt_pct <= 0:
                sum_opt_pct = 100.0

            seg_by_key = {s["key"]: s for s in segs_raw if isinstance(s, dict) and "key" in s} if segs_raw else {}

            # ── 2단계: 세그먼트 점수 및 감점 여부 사전 계산 ──
            computed_seg_pts = {}
            sum_earned_pt = 0.0
            has_valid_segs = False

            for key in seg_keys:
                seg = seg_by_key.get(key)
                brand_opt_pct = seg_opt_pct_map.get(key, 0.0)
                # 정규화된 만점 적용
                brand_seg_max_pt = m_weight * (brand_opt_pct / sum_opt_pct)

                if seg:
                    has_valid_segs = True
                    valM = float(seg.get("valM", 0.0))
                    targetM = float(seg.get("targetM", 0.0))
                    pct = float(seg.get("pct", 0.0))

                    if b.get("is_p1_aggregate"):
                        # P1 집계의 경우, seg에 이미 가중평균된 earned_pt와 is_over_120이 들어있음
                        earned_pt = min(float(seg.get("earned_pt", 0.0)), brand_seg_max_pt)
                        is_over_120 = bool(seg.get("is_over_120", False))
                    else:
                        real_pct = (valM / targetM * 100.0) if targetM > 0 else pct
                        is_over_120 = (real_pct > 120.0)
                        
                        # pct 최대 100으로 제한 후 연산, 세그먼트 만점 켜핑
                        capped_pct = min(pct, 100.0)
                        raw_earned_pt = brand_seg_max_pt * (capped_pct / 100.0)
                        earned_pt = raw_earned_pt * 0.9 if is_over_120 else raw_earned_pt
                        earned_pt = min(earned_pt, brand_seg_max_pt)
                    
                    computed_seg_pts[key] = {
                        "earned_pt": earned_pt,
                        "brand_seg_max_pt": brand_seg_max_pt,
                        "is_over_120": is_over_120,
                        "valM": valM,
                        "exists": True
                    }
                    sum_earned_pt += earned_pt
                else:
                    computed_seg_pts[key] = {
                        "earned_pt": 0.0,
                        "brand_seg_max_pt": brand_seg_max_pt,
                        "is_over_120": False,
                        "valM": 0.0,
                        "exists": False
                    }

            # ── 4단계: 1위 브랜드 기준 헤더에 세그먼트 만점을 정규화하여 표기 ──
            if "_ref_opt_pct_map" not in m_info:
                ref_opt_map = {}
                ref_sum_opt = 0.0
                if ref_b:
                    ref_sd = detail_data.get(ref_b.get("store", "")) or {}
                    ref_nd = ref_sd.get(ref_b.get("name", "")) or {}
                    ref_det = ref_nd.get(ref_b.get("type_label", "")) or {}
                    ref_mdata = ref_det.get(m_id) or {}
                    for s in (ref_mdata.get("segs") or []):
                        if isinstance(s, dict) and "key" in s:
                            op_val = float(s.get("opt_pct", 0.0))
                            ref_opt_map[s["key"]] = op_val
                            ref_sum_opt += op_val
                for ab in sorted_brands:
                    ab_sd = detail_data.get(ab.get("store", "")) or {}
                    ab_nd = ab_sd.get(ab.get("name", "")) or {}
                    ab_det = ab_nd.get(ab.get("type_label", "")) or {}
                    ab_mdata = ab_det.get(m_id) or {}
                    for s in (ab_mdata.get("segs") or []):
                        k = s.get("key")
                        if k and k not in ref_opt_map:
                            op_val = float(s.get("opt_pct", 0.0))
                            ref_opt_map[k] = op_val
                            ref_sum_opt += op_val
                if ref_sum_opt <= 0:
                    ref_sum_opt = 100.0
                m_info["_ref_opt_pct_map"] = ref_opt_map
                m_info["_ref_sum_opt_pct"] = ref_sum_opt

                # [v175] 1위 브랜드 기준의 각 세그먼트 만점을 계산하여 저장 (모든 브랜드 컬럼 캡핑 통일용)
                ref_seg_max_map = {}
                ref_m_weight = ref_weights.get(w_key, m_weight)
                for k, op_val in ref_opt_map.items():
                    ref_seg_max_map[k] = ref_m_weight * (op_val / ref_sum_opt)
                m_info["_ref_seg_max_pt_map"] = ref_seg_max_map

            ref_opt_pct_map = m_info["_ref_opt_pct_map"]
            ref_sum_opt_pct = m_info["_ref_sum_opt_pct"]
            ref_seg_max_pt_map = m_info["_ref_seg_max_pt_map"]

            # ── 3단계: 지표 합계 칸 쓰기 (ref_seg_max_pt 기준 캡핑 합산) ──
            if has_valid_segs:
                # ref_seg_max_pt 기준으로 각 세그먼트 earned_pt를 캡핑 후 합산
                capped_sum = 0.0
                for key in seg_keys:
                    c = computed_seg_pts.get(key)
                    if c and c["exists"]:
                        r_max = ref_seg_max_pt_map.get(key, 0.0)
                        d_max = r_max if r_max > 0 else c["brand_seg_max_pt"]
                        capped_sum += min(c["earned_pt"], d_max) if d_max > 0 else c["earned_pt"]
                raw_earned = min(capped_sum, global_m_weight)
                if score_mode == "100_percent" and global_m_weight > 0:
                    earned_score = (raw_earned / global_m_weight) * 100.0
                else:
                    earned_score = raw_earned
            else:
                raw_score_0_to_100 = float(b.get(m_id, 0.0))
                if score_mode == "100_percent":
                    earned_score = raw_score_0_to_100
                else:
                    earned_score = min(raw_score_0_to_100 * (global_m_weight / 100.0), global_m_weight)

            earned_score_rounded = _round1(earned_score)
            display_max_weight = 100.0 if score_mode == "100_percent" else global_m_weight
            
            if is_sales_execution:
                score_str = _fmt_num(earned_score_rounded)
            else:
                score_str = _fmt_score(earned_score_rounded)
                
            styles.apply_score_cell(
                ws.cell(row=row_idx, column=col_offset),
                score_str, earned_score_rounded, display_max_weight,
            )
            
            if is_sales_execution:
                fill = styles.fill_for_score(earned_score_rounded, display_max_weight)
                if fill is styles.fill_red:
                    tot_target = 0.0
                    tot_val = 0.0
                    for k in seg_keys:
                        c_info = computed_seg_pts.get(k)
                        if c_info and c_info['exists']:
                            tot_target += float(seg_by_key.get(k, {}).get('targetM', 0.0))
                            tot_val += float(c_info['valM'])
                    shortage = tot_val - tot_target
                    if shortage < 0:
                        metric_title = m_info['title']
                        red_lights.append(f"{metric_title} {score_str} ({int(shortage)}백)")
            if row_idx == ROW_DATA:
                # 4행 합계 컬럼: 정상/상설 만점을 각각 표기
                if score_mode == "100_percent":
                    _h3_total_val = "100점"
                else:
                    _nw, _ow = metric_norm_outl_w.get(m_id, (0, 0))
                    if _nw != _ow and _nw > 0 and _ow > 0:
                        _h3_total_val = f"정상{_nw}점\n상설{_ow}점"
                    elif global_m_weight > 0:
                        _h3_total_val = f"{int(round(global_m_weight))}점"
                    else:
                        _h3_total_val = "-"
                styles.apply_header_cell(
                    ws.cell(row=ROW_H3, column=col_offset), _h3_total_val
                )
            col_offset += 1

            if not segs_raw:
                for key in seg_keys:
                    styles.apply_yellow_cell(ws.cell(row=row_idx, column=col_offset), empty_txt)
                    if row_idx == ROW_DATA:
                        if score_mode == "100_percent":
                            _num_segs = len(seg_keys) if seg_keys else 1
                            _h3_seg_val = f"{int(round(100.0 / _num_segs))}점"
                        else:
                            ref_seg_max_pt = ref_seg_max_pt_map.get(key, 0.0)
                            _nw, _ow = metric_norm_outl_w.get(m_id, (0, 0))
                            _num_segs = len(seg_keys) if seg_keys else 1
                            if _nw != _ow and _nw > 0 and _ow > 0:
                                _seg_n = round(_nw / _num_segs) if _num_segs > 0 else _nw
                                _seg_o = round(_ow / _num_segs) if _num_segs > 0 else _ow
                                _h3_seg_val = f"정상{_seg_n}점\n상설{_seg_o}점"
                            else:
                                _h3_seg_val = f"{int(round(ref_seg_max_pt))}점" if ref_seg_max_pt > 0 else "-"
                        styles.apply_header_cell(
                            ws.cell(row=ROW_H3, column=col_offset), _h3_seg_val
                        )
                    col_offset += 1
                continue

            for key in seg_keys:
                c_info = computed_seg_pts[key]
                brand_seg_max_pt = c_info["brand_seg_max_pt"]
                ref_seg_max_pt = ref_seg_max_pt_map.get(key, 0.0)

                if c_info["exists"]:
                    valM = c_info["valM"]
                    is_over_120 = c_info["is_over_120"]

                    # [v174] 표시 점수는 헤더 만점(ref_seg_max_pt)과 브랜드 만점 중 작은 값으로 상한 캡핑
                    # → 헤더에 15점이라고 표시되면 절대 15점을 초과할 수 없음
                    display_max = ref_seg_max_pt if ref_seg_max_pt > 0 else brand_seg_max_pt
                    earned_pt = min(c_info["earned_pt"], display_max) if display_max > 0 else c_info["earned_pt"]

                    if score_mode == "100_percent":
                        if global_m_weight > 0:
                            # 100점 만점 모드에서의 세그먼트 환산 점수
                            # 100점 스케일 하에서 세그먼트의 비중 = ref_seg_max_pt / global_m_weight * 100
                            # 따라서 earned_pt 역시 (earned_pt / global_m_weight) * 100 으로 스케일링
                            earned_pt = (earned_pt / global_m_weight) * 100.0
                            display_max = (display_max / global_m_weight) * 100.0
                        else:
                            earned_pt = 0.0
                            display_max = 0.0

                    if brand_seg_max_pt <= 0:
                        txt = f"{_fmt_num(valM)}\n(-)" if valM > 0 else empty_txt
                        styles.apply_yellow_cell(ws.cell(row=row_idx, column=col_offset), txt)
                    else:
                        cell = ws.cell(row=row_idx, column=col_offset)
                        valM_r = _round1(valM)
                        earned_pt_r = _round1(earned_pt)
                        if valM_r == 0 and earned_pt_r == 0:
                            cell.value = "0점"
                        else:
                            cell.value = f"{_fmt_num(valM)}\n({_fmt_score(earned_pt)})"
                        cell.alignment = styles.align_center
                        cell.border = styles.border_thin
                        
                        if is_over_120:
                            cell.fill = styles.fill_yellow
                            cell.font = styles.font_black
                        else:
                            # 색상 판정도 display_max 기준으로 통일
                            fill = styles.fill_for_score(earned_pt, display_max)
                            cell.fill = fill
                            cell.font = styles.font_for_fill(fill)

                    if row_idx == ROW_DATA:
                        # 4행 세그먼트 컬럼: 정상/상설 비례 만점을 각각 표기
                        if score_mode == "100_percent":
                            _num_segs = len(seg_keys) if seg_keys else 1
                            _ref_sum = sum(ref_seg_max_pt_map.get(k, 0.0) for k in seg_keys) if seg_keys else 0.0
                            if _ref_sum > 0:
                                _seg_ratio = ref_seg_max_pt / _ref_sum
                                _h3_seg_val = f"{int(round(_seg_ratio * 100.0))}점"
                            else:
                                _h3_seg_val = f"{int(round(100.0 / _num_segs))}점"
                        else:
                            _nw, _ow = metric_norm_outl_w.get(m_id, (0, 0))
                            _num_segs = len(seg_keys) if seg_keys else 1
                            _ref_sum = sum(ref_seg_max_pt_map.get(k, 0.0) for k in seg_keys) if seg_keys else 0.0
                            if _nw != _ow and _nw > 0 and _ow > 0 and _ref_sum > 0:
                                _seg_ratio = ref_seg_max_pt / _ref_sum if _ref_sum > 0 else 0.0
                                _seg_n = round(_nw * _seg_ratio)
                                _seg_o = round(_ow * _seg_ratio)
                                _h3_seg_val = f"상설{_seg_o}점\n정상{_seg_n}점"
                            else:
                                _h3_seg_val = f"{int(round(ref_seg_max_pt))}점" if ref_seg_max_pt > 0 else "-"
                        styles.apply_header_cell(
                            ws.cell(row=ROW_H3, column=col_offset), _h3_seg_val
                        )
                else:
                    styles.apply_yellow_cell(ws.cell(row=row_idx, column=col_offset), empty_txt)
                    if row_idx == ROW_DATA:
                        _nw, _ow = metric_norm_outl_w.get(m_id, (0, 0))
                        _num_segs = len(seg_keys) if seg_keys else 1
                        _ref_sum = sum(ref_seg_max_pt_map.get(k, 0.0) for k in seg_keys) if seg_keys else 0.0
                        if _nw != _ow and _nw > 0 and _ow > 0 and _ref_sum > 0:
                            _seg_ratio = ref_seg_max_pt / _ref_sum if _ref_sum > 0 else 0.0
                            _seg_n = round(_nw * _seg_ratio)
                            _seg_o = round(_ow * _seg_ratio)
                            _h3_seg_val = f"상설{_seg_o}점\n정상{_seg_n}점"
                        else:
                            _h3_seg_val = f"{int(round(ref_seg_max_pt))}점" if ref_seg_max_pt > 0 else "-"
                        styles.apply_header_cell(
                            ws.cell(row=ROW_H3, column=col_offset), _h3_seg_val
                        )
                col_offset += 1

        if is_sales_execution:
            txt = "\n".join(red_lights)
            cell = ws.cell(row=row_idx, column=col_offset)
            cell.value = txt
            cell.alignment = styles.align_center
            cell.border = styles.border_thin
            col_offset += 1

        # [v177] 복사붙여넣기 시 PPT 슬라이드 비율에 맞춘 최적의 행높이(45)로 증대
        ws.row_dimensions[row_idx].height = 45
        row_idx += 1

    # Data row heights
    for r in range(ROW_DATA, row_idx):
        ws.row_dimensions[r].height = 25
        
    # [v177] 복사붙여넣기 시 PPT 슬라이드 비율에 맞춘 최적의 열너비 자동 분배
    for col_i in range(1, total_cols + 1):
        col_letter = get_column_letter(col_i)
        val_h1 = ws.cell(row=ROW_H1, column=col_i).value
        
        val_str = str(val_h1 or "").strip()
        if "랭크" in val_str:
            ws.column_dimensions[col_letter].width = 4.5
        elif "복종" in val_str:
            ws.column_dimensions[col_letter].width = 5.5
        elif "지점" in val_str:
            ws.column_dimensions[col_letter].width = 7.0
        elif "브랜드" in val_str:
            ws.column_dimensions[col_letter].width = 10.0
        elif "총" in val_str and "점수" in val_str:
            ws.column_dimensions[col_letter].width = 8.0
        elif "총보유" in val_str or "재고액" in val_str:
            ws.column_dimensions[col_letter].width = 10.0
        else:
            # 할인율, BEST상품, 신선도, 시즌, 아이템 세그먼트 셀들은 PPT 균형에 맞춰 균일 분배
            ws.column_dimensions[col_letter].width = 7.5


def export_all_in_one_excel_bytes(data: dict):
    """지점별 시트 멀티 워크북 — 전체 카테고리·전체 지표 통합."""
    if not data or "error" in data:
        return None
    brands = data.get("BRANDS", [])
    detail_data = data.get("DETAIL", {})
    if not brands:
        return None

    stores = data.get("STORES") or sorted({b.get("store") for b in brands if b.get("store")})
    wb = Workbook()
    wb.remove(wb.active)
    used_names: set = set()

    for store in stores:
        store_brands = [b for b in brands if b.get("store") == store]
        if not store_brands:
            continue
        ws = wb.create_sheet(title=_sanitize_sheet_name(store, used_names))
        _fill_exposure_sheet(
            ws, store_brands, detail_data,
            title_store=store, title_cat="전체 카테고리",
            metrics_filter=ALL_METRICS, show_store_col=False,
        )

    if not wb.sheetnames:
        return None
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_to_excel_bytes(
    data=None,
    store_filter="전체 지점",
    cat_filter="전체 카테고리",
    metrics_filter=None,
    score_mode: str = "weighted",
    is_sales_execution: bool = False,
):
    """하위 호환 — 단일 시트."""
    if metrics_filter is None:
        metrics_filter = ALL_METRICS
    if not data:
        data = load_dashboard_data()
        if "error" in data:
            return None
    brands = data.get("BRANDS", [])
    detail_data = data.get("DETAIL", {})
    filtered = [
        b for b in brands
        if (store_filter in ("전체 지점", "전체") or b.get("store") == store_filter)
        and (cat_filter in ("전체 카테고리", "전체") or b.get("category") == cat_filter)
    ]
    if not filtered:
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_title_from_filters(store_filter, cat_filter)
    _fill_exposure_sheet(
        ws, filtered, detail_data,
        title_store=store_filter, title_cat=cat_filter,
        metrics_filter=metrics_filter,
        show_store_col=(store_filter == "전체 지점"),
        score_mode=score_mode,
        is_sales_execution=is_sales_execution,
    )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_p1_summary_excel_bytes(data: dict, cat_filter: str, metrics_filter=None, score_mode: str = "weighted"):
    """[v173] P1 대시보드 다운로드: 카테고리 필터 기반 지점별 상세 지표 집계"""
    if metrics_filter is None:
        metrics_filter = ALL_METRICS
    if not data or "error" in data:
        return None
        
    stores = data.get("STORES", [])
    brands = data.get("BRANDS", [])
    detail = data.get("DETAIL", {})
    
    if not stores or not brands:
        return None

    metrics = ["dis", "best", "fresh", "season", "item"]
    agg_brands = []
    agg_detail = {}

    for store in stores:
        # 해당 지점/카테고리의 브랜드 목록
        if cat_filter in ("전체 카테고리", "전체"):
            store_brands = [b for b in brands if b.get("store") == store]
        else:
            store_brands = [b for b in brands if b.get("store") == store and b.get("category") == cat_filter]
        if not store_brands:
            continue

        b_count = len(store_brands)
        agg_tM = sum(float(b.get("tM", 0.0)) for b in store_brands)
        agg_sM = sum(float(b.get("sM", 0.0)) for b in store_brands)
        
        rep_b = store_brands[0]
        agg_b = dict(rep_b)
        agg_b["name"] = "지점 합계"
        agg_b["type_label"] = "전체"
        agg_b["sM"] = agg_sM
        agg_b["tM"] = agg_tM
        agg_b["is_p1_aggregate"] = True  # P1 집계 브랜드 플래그 추가

        # 개별 지표 점수(0~100점 스케일)도 가중평균하여 agg_b에 저장
        for m in metrics:
            if agg_tM > 0:
                agg_m_val = sum(float(b.get(m, 0.0)) * (float(b.get("tM", 0.0)) / agg_tM) for b in store_brands)
            else:
                agg_m_val = sum(float(b.get(m, 0.0)) for b in store_brands) / b_count if b_count > 0 else 0.0
            agg_b[m] = agg_m_val

        if store not in agg_detail:
            agg_detail[store] = {}
        agg_detail[store]["지점 합계"] = {"전체": {}}
        store_det = agg_detail[store]["지점 합계"]["전체"]

        # 5개 지표 점수의 합이 총점수가 되도록 실시간 합산
        agg_calc = 0.0
        agg_max_weights = 0.0
        for m in metrics:
            store_det[m] = {"segs": []}
            seg_map = {}
            
            # 가중평균 가중치 합산
            sum_weights = sum(float(b.get("tM", 0.0)) for b in store_brands)
            
            for b in store_brands:
                b_name = b.get("name")
                b_type = b.get("type_label")
                b_tM = float(b.get("tM", 0.0))
                
                # 브랜드별 가중치 (tM 비중, 없으면 단순평균)
                b_weight = (b_tM / sum_weights) if sum_weights > 0 else (1.0 / len(store_brands))
                
                s_weights = b.get("scoring_guide", {}).get("score_weights", {})
                w_key = "sea" if m == "season" else m
                m_weight = s_weights.get(w_key, 0.0)
                
                m_data = detail.get(store, {}).get(b_name, {}).get(b_type, {}).get(m, {})
                segs_raw = m_data.get("segs", [])
                
                # opt_pct 합계 계산
                seg_opt_pct_map = {}
                sum_opt_pct = 0.0
                for s in segs_raw:
                    if isinstance(s, dict) and "key" in s:
                        op_val = float(s.get("opt_pct", 0.0))
                        seg_opt_pct_map[s["key"]] = op_val
                        sum_opt_pct += op_val
                if sum_opt_pct <= 0:
                    sum_opt_pct = 100.0
                    
                for s in segs_raw:
                    if not isinstance(s, dict) or "key" not in s:
                        continue
                    k = s["key"]
                    brand_opt_pct = seg_opt_pct_map.get(k, 0.0)
                    brand_seg_max_pt = m_weight * (brand_opt_pct / sum_opt_pct)
                    
                    valM = float(s.get("valM", 0.0))
                    targetM = float(s.get("targetM", 0.0))
                    pct = float(s.get("pct", 0.0))
                    
                    real_pct = (valM / targetM * 100.0) if targetM > 0 else pct
                    is_over_120 = (real_pct > 120.0)
                    
                    # pct 켜핑 100으로, 세그먼트 만점 켜핑 적용
                    capped_pct = min(pct, 100.0)
                    raw_earned_pt = brand_seg_max_pt * (capped_pct / 100.0)
                    earned_pt = raw_earned_pt * 0.9 if is_over_120 else raw_earned_pt
                    earned_pt = min(earned_pt, brand_seg_max_pt)
                    
                    if k not in seg_map:
                        seg_map[k] = {
                            "valM": 0.0,
                            "targetM": 0.0,
                            "opt_pct": float(s.get("opt_pct", 0.0)),
                            "earned_pt": 0.0,
                            "l": s.get("l"),
                            "label": s.get("label")
                        }
                    seg_map[k]["valM"] += valM
                    seg_map[k]["targetM"] += targetM
                    seg_map[k]["earned_pt"] += earned_pt * b_weight

            # 세그먼트 합산 후 지표 만점으로 켜핑 후 총점에 누적
            m_total_earned = 0.0
            # m_weight: 각 브랜드의 tM 가중평균으로 지표 만점 계산
            avg_m_weight = (
                sum(b.get("scoring_guide", {}).get("score_weights", {}).get("sea" if m == "season" else m, 0.0) * (float(b.get("tM", 0.0)) / sum_weights)
                    for b in store_brands) if sum_weights > 0 else
                sum(b.get("scoring_guide", {}).get("score_weights", {}).get("sea" if m == "season" else m, 0.0)
                    for b in store_brands) / len(store_brands)
            )
            for k, v in seg_map.items():
                agg_real_pct = (v["valM"] / v["targetM"] * 100.0) if v["targetM"] > 0 else 0.0
                agg_is_over_120 = (agg_real_pct > 120.0)
                m_total_earned += v["earned_pt"]
                store_det[m]["segs"].append({
                    "key": k,
                    "valM": v["valM"],
                    "targetM": v["targetM"],
                    "opt_pct": v["opt_pct"],
                    "pct": agg_real_pct,
                    "earned_pt": v["earned_pt"],
                    "is_over_120": agg_is_over_120,
                    "l": v.get("l"),
                    "label": v.get("label")
                })
            # 지표 만점(avg_m_weight) 컵핑 후 아웃터 루프에서 총점 누적
            m_earned = min(m_total_earned, avg_m_weight)
            
            metric_filter_map = {"dis": "할인율", "best": "BEST상품", "fresh": "신선도", "season": "시즌", "item": "아이템"}
            is_all = ("전체" in metrics_filter) or (len(metrics_filter) >= 4)

            if is_all or metric_filter_map[m] in metrics_filter:
                agg_calc += m_earned
                agg_max_weights += avg_m_weight

        # 선택된 지표에 대해 100점 만점으로 환산 (정규화)
        if score_mode == "100_percent":
            selected_count = sum(1 for m in metrics if (is_all or metric_filter_map[m] in metrics_filter) and avg_m_weight > 0)
            if selected_count > 0:
                normalized_calc = (agg_calc / agg_max_weights) * 100.0 if agg_max_weights > 0 else 0.0
            else:
                normalized_calc = 0.0
        else:
            if agg_max_weights > 0 and agg_max_weights < 100.0 and not (("전체" in metrics_filter) or (len(metrics_filter) >= 4)):
                normalized_calc = (agg_calc / agg_max_weights) * 100.0
            else:
                normalized_calc = agg_calc
            
        agg_b["calculated_total"] = _round1(min(normalized_calc, 100.0))
        agg_brands.append(agg_b)

    if not agg_brands:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = _sanitize_sheet_name(f"지점요약_{cat_filter}", set())
    
    _fill_exposure_sheet(
        ws, agg_brands, agg_detail,
        title_store="전체 지점", title_cat=cat_filter,
        metrics_filter=metrics_filter,
        show_store_col=True,
        hide_brand_col=True,
        score_mode=score_mode,
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _apply_dashboard_score_cell(cell, score, styles):
    """대시보드 100점 기준 셀: 80점↑ 녹색, 60점↑ 노란색, 미만 빨간색."""
    cell.value = f"{int(round(score))}점"
    if score >= 80:
        cell.fill = styles.fill_green
        cell.font = styles.font_white_bold
    elif score >= 60:
        cell.fill = styles.fill_yellow
        cell.font = styles.font_black
    else:
        cell.fill = styles.fill_red
        cell.font = styles.font_white_bold
    cell.alignment = styles.align_center
    cell.border = styles.border_thin


def export_p1_dashboard_excel_bytes(data: dict, cat_filter: str, metrics_filter=None):
    """100점 환산 기준 대시보드형 Excel: 세그먼트 없이 5개 지표를 각 100점 만점으로 표시."""
    if metrics_filter is None:
        metrics_filter = ALL_METRICS
    metrics_filter = _normalize_metrics_filter(metrics_filter)

    if not data or "error" in data:
        return None

    stores = data.get("STORES", [])
    brands = data.get("BRANDS", [])
    if not stores or not brands:
        return None

    METRIC_KEYS = [
        ("dis",    "할인율"),
        ("best",   "BEST상품"),
        ("fresh",  "신선도"),
        ("season", "시즌"),
    ]
    active_metrics = [(k, l) for k, l in METRIC_KEYS if l in metrics_filter]

    agg_rows = []
    for store in stores:
        if cat_filter in ("전체 카테고리", "전체"):
            sb = [b for b in brands if b.get("store") == store]
        else:
            sb = [b for b in brands if b.get("store") == store and b.get("category") == cat_filter]
        if not sb:
            continue

        b_count = len(sb)
        agg_tM = sum(float(b.get("tM", 0.0)) for b in sb)
        cat_display = "전체" if cat_filter == "전체 카테고리" else cat_filter

        row = {"store": store, "category": cat_display}
        for m_key, _ in METRIC_KEYS:
            if agg_tM > 0:
                row[m_key] = min(100.0, sum(
                    float(b.get(m_key, 0.0)) * (float(b.get("tM", 0.0)) / agg_tM) for b in sb
                ))
            else:
                row[m_key] = min(100.0, sum(float(b.get(m_key, 0.0)) for b in sb) / b_count)

        sel_scores = [row[mk] for mk, ml in METRIC_KEYS if ml in metrics_filter]
        row["total"] = sum(sel_scores) / len(sel_scores) if sel_scores else 0.0
        agg_rows.append(row)

    if not agg_rows:
        return None

    agg_rows.sort(key=lambda x: x["total"], reverse=True)

    styles = _STYLES
    wb = Workbook()
    ws = wb.active
    ws.title = _sanitize_sheet_name(f"대시보드_{cat_filter}", set())

    col_defs = [("랭크", 5), ("복종", 10), ("지점", 12), ("총점\n(100점)", 9)]
    for _, m_label in active_metrics:
        col_defs.append((f"{m_label}\n(100점)", 9))

    for ci, (_, w) in enumerate(col_defs, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    total_cols = len(col_defs)

    title_cell = ws.cell(row=1, column=1,
        value=f"[한국유통] 상품구색 노출 100점 환산 대시보드 ({cat_filter})")
    title_cell.font = styles.font_title
    title_cell.alignment = styles.align_center
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.row_dimensions[1].height = 25

    for ci, (hdr, _) in enumerate(col_defs, 1):
        styles.apply_header_cell(ws.cell(row=2, column=ci), value=hdr)
    ws.row_dimensions[2].height = 32

    for rank, row in enumerate(agg_rows, 1):
        ri = 2 + rank
        ws.row_dimensions[ri].height = 18
        styles.apply_plain_cell(ws.cell(row=ri, column=1), rank)
        styles.apply_plain_cell(ws.cell(row=ri, column=2), row["category"])
        styles.apply_plain_cell(ws.cell(row=ri, column=3), row["store"])
        col = 4
        _apply_dashboard_score_cell(ws.cell(row=ri, column=col), row["total"], styles)
        col += 1
        for m_key, _ in active_metrics:
            _apply_dashboard_score_cell(ws.cell(row=ri, column=col), row[m_key], styles)
            col += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_p2_dashboard_excel_bytes(data: dict, store_filter: str, cat_filter: str, metrics_filter=None):
    """P2 브랜드 상세 - 100점 환산 대시보드 Excel: 브랜드별 5개 지표를 각 100점 만점으로 표시."""
    if metrics_filter is None:
        metrics_filter = ALL_METRICS
    metrics_filter = _normalize_metrics_filter(metrics_filter)

    if not data or "error" in data:
        return None

    brands = data.get("BRANDS", [])
    filtered = [
        b for b in brands
        if (store_filter in ("전체 지점", "전체") or b.get("store") == store_filter)
        and (cat_filter in ("전체 카테고리", "전체") or b.get("category") == cat_filter)
    ]
    if not filtered:
        return None

    METRIC_KEYS = [
        ("dis",    "할인율"),
        ("best",   "BEST상품"),
        ("fresh",  "신선도"),
        ("season", "시즌"),
    ]
    active_metrics = [(k, l) for k, l in METRIC_KEYS if l in metrics_filter]

    def _row_total(b):
        scores = [min(100.0, float(b.get(mk, 0.0))) for mk, ml in METRIC_KEYS if ml in metrics_filter]
        return sum(scores) / len(scores) if scores else 0.0

    rows = sorted(filtered, key=_row_total, reverse=True)

    styles = _STYLES
    wb = Workbook()
    ws = wb.active
    ws.title = _sanitize_sheet_name(f"브랜드대시보드_{store_filter}", set())

    col_defs = [("랭크", 5), ("복종", 10), ("지점", 12), ("브랜드", 12), ("총점\n(100점)", 9)]
    for _, m_label in active_metrics:
        col_defs.append((f"{m_label}\n(100점)", 9))

    for ci, (_, w) in enumerate(col_defs, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    total_cols = len(col_defs)

    title_cell = ws.cell(row=1, column=1,
        value=f"[한국유통] 브랜드 100점 환산 대시보드 ({store_filter} / {cat_filter})")
    title_cell.font = styles.font_title
    title_cell.alignment = styles.align_center
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.row_dimensions[1].height = 25

    for ci, (hdr, _) in enumerate(col_defs, 1):
        styles.apply_header_cell(ws.cell(row=2, column=ci), value=hdr)
    ws.row_dimensions[2].height = 32

    for rank, b in enumerate(rows, 1):
        ri = 2 + rank
        ws.row_dimensions[ri].height = 18
        styles.apply_plain_cell(ws.cell(row=ri, column=1), rank)
        styles.apply_plain_cell(ws.cell(row=ri, column=2), b.get("category", ""))
        styles.apply_plain_cell(ws.cell(row=ri, column=3), b.get("store", ""))
        styles.apply_plain_cell(ws.cell(row=ri, column=4), b.get("name", ""))
        col = 5
        _apply_dashboard_score_cell(ws.cell(row=ri, column=col), _row_total(b), styles)
        col += 1
        for m_key, _ in active_metrics:
            score = min(100.0, float(b.get(m_key, 0.0)))
            _apply_dashboard_score_cell(ws.cell(row=ri, column=col), score, styles)
            col += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
